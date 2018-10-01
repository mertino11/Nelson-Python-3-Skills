# NHTCU WebScanner VERSION 1.0.0

# GUIDES FOLLOWED or QUESTIONS ASKED
# 1) https://stackoverflow.com/questions/42519899/create-ip-and-port-scanner-python
# 2) http://www.coderholic.com/python-port-scanner/
# 3) http://www.opentechguides.com/how-to/article/python/57/python-ping-subnet.html
# 4) https://stackoverflow.com/questions/46195221/formatting-datetime-output-in-python
# 5) https://stackoverflow.com/questions/35725732/what-is-the-function-of-sock-stream
# 6) https://docs.python.org/2/library/subprocess.html
# 7) https://goo.gl/Huz9M1
# 8) https://stackoverflow.com/questions/47778622/how-can-i-write-data-in-a-loop-in-openpyxl
# 9) http://naelshiab.com/tutorial-send-email-python/


# MODULES / PACKAGES

import ctypes
import hashlib
import ipaddress
import smtplib
import socks
import subprocess
import subprocess
import time
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from netaddr import *
from openpyxl import *
from socket import *
from sys import platform

from backend.scripts.Csv import *


# Class for the fancy colors
class color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

# Local date and time
localtime = time.localtime(time.time())
now = datetime.now()

# Start time of scan
t1 = datetime.now()

# Counts ports that are open.
portcount = 0

# Configure subprocess to hide the console window
# proc = subprocess.Popen(["echo","hello world"],stdout=subprocess.PIPE)

# print (proc.communicate())

# Name of the scan.
scanname = "Testscan1"

# Creates an XLSX Excel File
print(f"{color.DARKCYAN}Creating a report... {color.END}")
book = Workbook()
sheet = book.active
book.save("Scanreport.xlsx")

# Scanning is running from which host (could be useful to log who and when runs this script).
print(color.BLUE)
print("User ", '"', gethostname(), '"', " is now running a scan at the following date and time:",
      datetime.today() \
      .strftime("%A, %d %B %Y at {}:{}".format(now.hour, now.minute)), color.END, "\n")


def scanning(all_hosts, all_ports):
    portcount = 0  # Counts ports that are open.
    start_row = 1  # Adds a new row (to write HIT information) when it founds a HIT.
    t1 = datetime.now()  # Start time of scan

    # Loop where the IP-address is being pinged.
    for host in all_hosts:
        ping = subprocess.call(['ping', '-c', '3', str(host)])

        if ping == 0:
            print(f"{color.GREEN}")
            print(str(host), "is ONLINE!", color.END, "\n")
            print(f"{color.DARKCYAN}Trying to resolve hostname of {host}... {color.END} \n")

            # Tries to do a Reverse DNS Lookup (asking for a hostname by giving its IP-address).
            try:
                hostname = gethostbyaddr(str(host))
                print(f"{color.GREEN}Hostname FOUND! {color.END}\n")
                print(f"The hostname of {host} is {hostname}. \n")
                print(f"{color.DARKCYAN}Scanning ports in {hostname} ... {color.END} \n")

            except ImportError:
                hostname = 'NONE'
                print(f"{color.RED}Could not resolve hostname of {host}. {color.END}\n")

            # Loop where it scan ports within a range.
            for port in all_ports:
                print("\n")
                print(f"{color.DARKCYAN}Scanning Port: {port} ... {color.END}")

                # Creates an IPv4 socket + TCP Socket.
                # s = socket(AF_INET, SOCK_DGRAM) # UDP Socket
                s = socket(AF_INET, SOCK_STREAM)  # TCP Socket

                result = s.connect_ex((str(host), port))

                # If port is open...
                if (result == 0):
                    print(f"{color.GREEN}Port {port} is OPEN! {color.END}\n")
                    portcount += 1

                    rule1 = True  # Rule 1 (one of the criteria that was written in the assignment paper)

                    # Trying to grab the banner of the port.
                    print(f"{color.DARKCYAN}Trying to grab banner from {hostname} / {host} ... {color.END}\n")
                    s.connect_ex((str(host), port))
                    s.send(b'GET / HTTP/1.1\n\n\n')
                    banner = s.recv(1024)

                    # If else statement if there is a match between the user's keyword and the banner.
                    if banner != 0:
                        print(f"{color.GREEN}Banner found!\n")

                        keyword = 'a'  # Keyword = What keyword the user wants to find in banner.

                        print(f"{color.DARKCYAN}Matching keyword with the banner... {color.END}\n")

                        if keyword in str(banner):
                            print(f"{color.GREEN}Keyword {keyword} is found in the banner! {color.END}\n")
                            rule2 = True  # Rule 2 (one of the criteria that was written in the assignment paper)
                            s.close()
                            print("HIT!\n")  # Each individual hit will send an e-mail to the recipient.
                            print(f"{color.DARKCYAN}Sending mail with HIT details... {color.END}\n")
                            sendHITMail(hostname, host, port, keyword, banner, scanname)

                            print(f"{color.DARKCYAN}Inputting HIT details in the report... {color.END}\n")

                            wb = load_workbook('Scanreport.xlsx')

                            hitdetails = (str(hostname), str(host), str(port), str(keyword), str(banner))
                            ws = wb.active
                            sheet = wb.get_sheet_by_name('Sheet')

                            sheet['A1'] = 'Hostname'
                            sheet['B1'] = 'IP'
                            sheet['C1'] = 'Port'
                            sheet['D1'] = 'Keyword'
                            sheet['E1'] = 'Banner'

                            start_row += 1
                            start_column = 1

                            for searchresult in hitdetails:
                                ws.cell(row=start_row, column=start_column).value = searchresult
                                start_column += 1

                            wb.save("scanreport.xlsx")

                        else:
                            print(f"{color.RED}Keyword was not found in banner. {color.END}\n")
                            rule2 = False  # Rule 2 (one of the criteria that was written in the assignment paper)
                            s.close()
                    else:
                        print(f"{color.RED}Failed to grab banner from {hostname} / {host}! {color.END}\n")
                        s.close()

                    print(f"{color.DARKCYAN}Converting banner to SHA1 hash... {color.END}")
                    sha1hash = hashlib.sha1(banner)
                    hex_dig = sha1hash.hexdigest()
                    print(f"SHA1: {hex_dig}")

                # If port is closed...
                else:
                    rule1 = False  # Rule 1 (one of the criteria that was written in the assignment paper)
                    s.close()

        elif ping == 2:
            print(f"{color.RED}No response from {host}. {color.END}\n")

        else:
            print(f"{color.RED}Ping to {host} has failed. {color.END}\n")

    # Check end time
    t2 = datetime.now()

    # Calculates the difference of time, to see how long it took to run the script.
    total = t2 - t1

    # Prints the end result of the scan.
    print("\n")
    print(f"{color.BLUE}______________________________________________{color.END}\n")
    print(f"{color.BLUE}Scanning completed in : {total} {color.END}\n")
    print(f"{color.BLUE}Ports open: {portcount} {color.END}\n")
    print(f"{color.BLUE}Scan time started:       {t1} {color.END}")
    print(f"{color.BLUE}Scan time ended:         {t2} {color.END}\n")
    print(f"{color.BLUE}______________________________________________{color.END}\n")
    sendMailReport(t1, t2, total, portcount, scanname)
    return portcount

def sendHITMail(host, port, hostname='', keyword='test',
                banner='udp heeft geen banner', scanname='wateenmooiename'):
    # Opening and reading text files.
    userMail = open("usermail.txt", "r")
    passwordMail = open("passwordmail.txt", "r")
    sendtoMail = open("sendtomail.txt", "r")

    # Email credentials from the sender
    email_user = (userMail.readline(20))
    email_password = (passwordMail.readline(18))
    email_sendto = (sendtoMail.readline(21))

    # MIMEPart - Part of the email
    msg = MIMEMultipart()
    msg['From'] = email_user
    msg['To'] = email_sendto
    msg['Subject'] = f'Individual hit on {host} on port {port} at {datetime.now():%A, %d %B %Y at %H:%M}'
    body = """\
         There is a hit on the following machine:
         Scan name: """ + str(scanname) + """
         Date and Time:  """ + datetime.today() \
        .strftime("%A, %d %B %Y at {}:{}".format(now.hour, now.minute)) + """ and """ + str(port) + """
         Hostname:  """ + str(hostname) + """
         IP-address:  """ + str(host) + """
         Port:  """ + str(port) + """
         Keyword:  """ + str(keyword) + """
         Banner: """ + """

         """ + str(banner) + """
                              """
    msg.attach(MIMEText(body, 'plain'))

    try:
        # Makes a secure TLS-connection to Gmail through port 587
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_user, email_password)

        # Sends the email to the recipient and closes the connection.
        text = msg.as_string()
        server.sendmail(email_user, email_sendto, text)
        server.quit()

        # Close the file that was open + read.
        userMail.close()
        passwordMail.close()

        print(f"Hit has been sent to {email_sendto}\n")

        sendtoMail.close()

    except:
        print(f"Hit cannot be sent to {email_sendto}\n")


def sendMailReport(t1, t2, total, portcount, scanname):
    # Opening and reading text files.
    userMail = open("usermail.txt", "r")
    passwordMail = open("passwordmail.txt", "r")
    sendtoMail = open("sendtomail.txt", "r")

    # Email credentials from the sender
    email_user = (userMail.readline(20))
    email_password = (passwordMail.readline(18))
    email_sendto = (sendtoMail.readline(21))

    # MIMEPart - Part of the email
    msg = MIMEMultipart()

    # Subject of the email
    msg['Subject'] = f'Your Scanning Report on {datetime.now():%A, %d %B %Y at %H:%M}'

    # Adding text inside the email (body)
    body = """\
         Your scanning report on """ + datetime.today() \
        .strftime("%A, %d %B %Y at {}:{}".format(now.hour, now.minute)) + """
         Scan name: """ + str(scanname) + """
         Scan time started:       """ + str(t1) + """
         Scan time ended:        """ + str(t2) + """
         Scan completed in:     """ + str(total) + """
         Ports open:  """ + str(portcount) + """
          """

    msg.attach(MIMEText(body, 'plain'))

    # Grabs a filename and open the filename
    filename = 'scanreport.xlsx'
    attachment = open(filename, 'rb')

    # Upload the file in MIMEBase as a base64 and reads the contents of the file.
    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)

    # Adds the header of the file, so the email knows what name of the filename is (so filename.txt)
    part.add_header('Content-Disposition', "attachment; filename= " + f"{filename}")

    # Attaches the document file as a text-string
    msg.attach(part)
    text = msg.as_string()
    try:
        # Makes a secure TLS-connection to Gmail through port 587
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_user, email_password)

        # Sends the email to the recipient and closes the connection.
        server.sendmail(email_user, email_sendto, text)
        server.quit()

        # Close the file that was open + read.
        userMail.close()
        passwordMail.close()

        print(f"Scan report has been sent to {email_sendto}\n")

        sendtoMail.close()

    except:
        print(f"Scan report cannot be sent to {email_sendto}\n")