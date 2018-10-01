# Test Environment that I made to create an XLSX file, write information it and then save it.

from openpyxl import *

book = Workbook()
sheet = book.active

sheet['A1'] = "Hostname"
sheet['B1'] = "IP"
sheet['C1'] = "Port"
sheet['D1'] = "Keyword"
sheet['E1'] = "Banner"
book.save("scanreport.xlsx")

# -----------------------------------------------

wb = load_workbook('scanreport.xlsx')

test1 = 'MERTINO11'
test2 = '192.168.2.10'
test3 = 'a'
test4 = 'Banner info'

nfc_east = (test1, test2,test3,test4)
wb = Workbook()
ws = wb.active

start_row = 2
start_column = 1

for team in nfc_east:
    ws.cell(row=start_row, column=start_column).value = team
    start_column += 1

start_row +=1
wb.save("scanreport.xlsx")
